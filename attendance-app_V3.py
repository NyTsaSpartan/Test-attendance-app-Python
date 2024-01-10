import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import shutil
import os
import csv
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook, load_workbook
from pandas import DataFrame, ExcelWriter


class RapportEnregistrementApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Rapport d'Enregistrement")

        # Configuration du style
        ttk.Style().configure("TButton", padding=6, relief="flat", background="#ccc")

        # Connexion à la base de données
        self.conn = None
        self.c = None

        # Zone gauche de l'interface
        self.setup_left_interface()

        # Zone centrale de l'interface
        self.setup_right_interface()

    def setup_left_interface(self):
        left_frame = tk.Frame(self.root)
        left_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nw")

        ttk.Label(left_frame, text="Convertir un fichier .csv puis importer une base de données").grid(row=0, column=0, pady=10, sticky="nw", columnspan=2)
        self.btn_importer_bd = ttk.Button(left_frame, text="Convertir un fichier\n       .csv en .db", command=self.call_csv_to_sql, width=20 )
        self.btn_importer_bd.grid(row=1, column=0, pady=5, sticky="w")

        self.btn_importer_bd = ttk.Button(left_frame, text="   Importer une\nbase de données", command=self.importer_base_de_donnees, width=20)
        self.btn_importer_bd.grid(row=1, column=1, pady=5, sticky="w")

        self.label_nom_bd_importee = ttk.Label(left_frame, text="BD importée : Aucune")
        self.label_nom_bd_importee.grid(row=2, column=0, pady=5, sticky="w")

        #ttk.Label(left_frame, text="Rechercher le nom d'une personne :").grid(row=3, column=0, pady=10, sticky="w")
        #self.saisie_personne = ttk.Combobox(left_frame, values=self.get_personnes(), width=30)
        #self.saisie_personne.grid(row=4, column=0, pady=5, sticky="w")

        #ttk.Label(left_frame, text="Sélectionnez une personne :").grid(row=5, column=0, pady=10, sticky="w")
        #self.personnes = ttk.Combobox(left_frame, values=self.get_personnes(), width=30)
        #self.personnes.grid(row=6, column=0, pady=5, sticky="w")

        ttk.Label(left_frame, text="Rechercher le nom d'une personne :").grid(row=3, column=0, pady=10, sticky="w")

        # Utiliser un Combobox pour la saisie et la sélection des personnes
        self.combobox_personnes = ttk.Combobox(left_frame, width=30)
        self.combobox_personnes.grid(row=4, column=0, pady=5, sticky="w")

        ttk.Label(left_frame, text="Choisissez une date disponible :").grid(row=7, column=0, pady=10, sticky="w")
        self.date_combobox = ttk.Combobox(left_frame, state="readonly")
        self.date_combobox.grid(row=8, column=0, pady=5, sticky="w")

        ttk.Button(left_frame, text="Exporter vers Excel", command=self.exporter_vers_excel, width=20).grid(row=9, column=0, pady=10, sticky="w")

        # Nouvelle étiquette pour afficher la somme des durées de présence à gauche
        self.label_total_heures_presence = tk.Label(left_frame, text="Total des heures de présence : ")
        self.label_total_heures_presence.grid(row=10, column=0, padx=0, sticky="nw", columnspan=2)

        # Nouvelle étiquette pour afficher la somme des durées de sortie à gauche
        #self.label_total_heures_sortie = tk.Label(left_frame, text="Total des heures de sortie : ")
        #self.label_total_heures_sortie.grid(row=11, column=0, padx=0, sticky="nw", columnspan=2)

        # Nouvelle étiquette pour afficher le nombre de sorties journalières à gauche
        self.label_nombre_sorties = tk.Label(left_frame, text="Nombre de sorties : ")
        self.label_nombre_sorties.grid(row=12, column=0, padx=0, sticky="w", columnspan=2)

        self.result_label = tk.Label(left_frame, text="")
        self.result_label.grid(row=13, column=0, padx=0, sticky="w", columnspan=2)

        ttk.Button(left_frame, text="Exporter Intervalles vers Excel", command=self.exporter_intervalles_vers_excel,
                   width=30).grid(row=14, column=0, pady=10, sticky="w")

        #self.personnes.bind("<<ComboboxSelected>>", self.mettre_a_jour_dates_disponibles)
        #self.date_combobox.bind("<<ComboboxSelected>>", self.afficher_rapports)
        #self.saisie_personne.bind("<KeyRelease>", self.mettre_a_jour_saisie_personne)
        #self.personnes.bind("<<ComboboxSelected>>", self.afficher_rapports)
        self.date_combobox.bind("<<ComboboxSelected>>", self.afficher_rapports)

        self.combobox_personnes.bind("<KeyRelease>", self.mettre_a_jour_combobox_personnes)
        self.combobox_personnes.bind("<<ComboboxSelected>>", self.afficher_rapports)
        self.combobox_personnes.bind("<<ComboboxSelected>>", self.mettre_a_jour_dates_disponibles)

    def setup_right_interface(self):
        center_frame = tk.Frame(self.root)
        center_frame.grid(row=0, column=1, padx=50, pady=20, sticky="nw")

        # Treeview pour afficher les intervalles de sortie
        self.treeview_intervalles_sortie = ttk.Treeview(center_frame, columns=("Durée", "Début", "Fin"))
        self.treeview_intervalles_sortie.heading("#0", text=" Heures de présence")
        self.treeview_intervalles_sortie.heading("Durée", text="Durée")
        self.treeview_intervalles_sortie.heading("Début", text="Début")
        self.treeview_intervalles_sortie.heading("Fin", text="Fin")
        self.treeview_intervalles_sortie.column(0, width=100)
        self.treeview_intervalles_sortie.column(1, width=100)
        self.treeview_intervalles_sortie.column(2, width=100)
        self.treeview_intervalles_sortie.grid(row=0, column=0, pady=10, sticky="nw")

        # Treeview pour afficher les rapports intermédiaires
        self.treeview_rapports_intermediaires = ttk.Treeview(center_frame, columns=("Durée", "Début", "Fin"))
        self.treeview_rapports_intermediaires.heading("#0", text="Temps de pauses")
        self.treeview_rapports_intermediaires.heading("Durée", text="Durée")
        self.treeview_rapports_intermediaires.heading("Début", text="Début")
        self.treeview_rapports_intermediaires.heading("Fin", text="Fin")
        self.treeview_rapports_intermediaires.column(0, width=100)
        self.treeview_rapports_intermediaires.column(1, width=100)
        self.treeview_rapports_intermediaires.column(2, width=100)
        self.treeview_rapports_intermediaires.grid(row=2, column=0, pady=10, sticky="sw")

    def call_csv_to_sql(self):
        csv_converter_dialog = tk.Toplevel(self.root)
        converter_instance = CSVtoDBConverter(csv_converter_dialog)
        csv_converter_dialog.protocol("WM_DELETE_WINDOW", lambda: self.close_csv_converter_dialog(csv_converter_dialog))

    def close_csv_converter_dialog(self, dialog):
        # Fermer la fenêtre parente (appelée lorsque la fenêtre est fermée)
        dialog.destroy()

    def importer_base_de_donnees(self):
        self.liste_personne = [] # liste des personnes dans la BD
        self.liste_date = [] # liste des dates dans la BD
        self.liste_HTT = [] # liste des heures totales de présence
        try:
            fichier_db = filedialog.askopenfilename(title="Sélectionner la base de données", filetypes=[("Fichiers de base de données", "*.db")])

            if fichier_db:
                # Copier la base de données sélectionnée dans le répertoire de travail
                shutil.copy(fichier_db, os.path.join(os.getcwd(), "imported_database.db"))

                # Mettre à jour le label avec le nom de la base de données importée
                self.label_nom_bd_importee.config(text=f"BD importée : {os.path.basename(fichier_db)}")

                # Connecter à la base de données copiée
                with sqlite3.connect('imported_database.db') as conn:
                    self.c = conn.cursor()

                # Mise à jour de la liste des personnes dans le menu déroulant
                #self.personnes['values'] = self.get_personnes()
                self.combobox_personnes['values'] = self.get_personnes()

                # Mettre à jour les dates disponibles dans la combobox
                self.mettre_a_jour_dates_disponibles()

        except Exception as e:
            print(f"Une erreur s'est produite lors de l'importation de la base de données : {e}")

    def mettre_a_jour_combobox_personnes(self, event=None):
        nom_saisi = self.combobox_personnes.get().lower()
        personnes = self.get_personnes()
        suggestions = [p for p in personnes if nom_saisi in p.lower()]

        # Mettre à jour les valeurs du combobox
        self.combobox_personnes['values'] = suggestions

    def get_personnes(self):
        if self.c:
            self.c.execute("SELECT DISTINCT nom_personne FROM pointeuse")
            personnes = [row[0] for row in self.c.fetchall()]
            return personnes
        else:
            return []

    def get_dates_disponibles(self):
        if self.c:
            self.c.execute(
                "SELECT DISTINCT strftime('%Y-%m-%d', heure) FROM pointeuse GROUP BY strftime('%Y-%m-%d', heure), nom_personne")
            dates_disponibles = [row[0] for row in self.c.fetchall()]
            return dates_disponibles
        else:
            return []

    def selectionner_personne(self):
        nom_saisi = self.saisie_personne.get()

        # Vous pouvez ajouter le code pour sélectionner la personne ici, par exemple, mettre à jour la combobox personnes.
        personnes = self.get_personnes()
        suggestions = [p for p in personnes if nom_saisi.lower() in p.lower()]

        if suggestions:
            self.personnes.set(suggestions[0])  # Sélectionnez la première suggestion

    def mettre_a_jour_combobox(self, combobox, values):
        combobox['values'] = values

    def mettre_a_jour_dates_disponibles(self, event=None):
        dates_disponibles = self.get_dates_disponibles()
        self.mettre_a_jour_combobox(self.date_combobox, dates_disponibles)
        self.afficher_rapports()

    def afficher_treeview_intervalles(self, treeview, intervalles):
        # Effacer le Treeview existant
        treeview.delete(*treeview.get_children())

        # Ajouter les intervalles au Treeview
        for i, (debut, fin) in enumerate(intervalles):
            debut_str = debut.strftime("%H:%M:%S")
            fin_str = fin.strftime("%H:%M:%S")
            duree = str(fin - debut)
            treeview.insert("", i, text=f"Heure de présence n°{i + 1}", values=(duree, debut_str, fin_str))

    def afficher_treeview_rapports_intermediaires(self, intervalles):
        # Effacer le Treeview existant
        for row in self.treeview_rapports_intermediaires.get_children():
            self.treeview_rapports_intermediaires.delete(row)

        # Ajouter les rapports intermédiaires au Treeview
        for i in range(len(intervalles) - 1):
            debut = intervalles[i][1].strftime("%H:%M:%S")
            fin = intervalles[i + 1][0].strftime("%H:%M:%S")
            duree = str(intervalles[i + 1][0] - intervalles[i][1])
            self.treeview_rapports_intermediaires.insert("", i, text=f"Sortie n°{i + 1}", values=(duree, debut, fin))

    def calculer_rapport(self, personne, date):
        try:
            if self.c:
                self.c.execute("SELECT heure FROM pointeuse WHERE nom_personne=? AND strftime('%Y-%m-%d', heure)=?",
                               (personne, date))
                heures_pointage = self.c.fetchall()

                if not heures_pointage:
                    return "Aucun rapport trouvé pour\n cette personne et cette date.", [], [], 0, "N/A", "N/A", "N/A"

                first_check = str(min(heures_pointage))
                first_check = first_check.replace("'", "")
                first_check = first_check.replace("(", "")
                first_check = first_check.replace(")", "")
                first_check = first_check.replace(",", "")

                last_check = str(max(heures_pointage))
                last_check = last_check.replace("'", "")
                last_check = last_check.replace("(", "")
                last_check = last_check.replace(")", "")
                last_check = last_check.replace(",", "")


                intervalles_presence, intervalles_sortie = self.calculer_intervalles(heures_pointage)
                if len(intervalles_sortie) > 0:
                    nombre_sorties = len(intervalles_sortie) - 1
                elif len(intervalles_sortie) == 0:
                    nombre_sorties = len(intervalles_sortie)

                # Calcul de la somme des durées de sortie
                total_heures_sortie = self.calculer_total_heures_sortie(intervalles_sortie)
                sum_intervalles_sortie = self.calculer_sum_intervalles_sortie(intervalles_sortie)

                sum_intervalles_presence = self.calculer_sum_intervalles_presence(intervalles_presence)
                total_heures_presence = self.calculer_total_heures_presence(intervalles_sortie)
                #print(sum_intervalles_sortie)

                rapport = f"Première Heure d'entrée :\n {first_check}\n\nDernière Heure de sortie :\n {last_check}\n\n"

                self.result_label.config(text=rapport)

                #if nombre_sorties == 0 and total_heures_presence > timedelta(hours=1) and total_heures_presence > timedelta(hours=9) and total_heures_sortie > timedelta(hours=1):
                #    total_heures_presence -= timedelta(hours=1)
                if (total_heures_presence > timedelta(hours=1)):
                    total_heures_presence -= timedelta(hours=1)

                return rapport, intervalles_presence, intervalles_sortie, nombre_sorties, sum_intervalles_presence, total_heures_presence, total_heures_sortie

            else:
                return "Aucune base de données importée. \n Veuillez importer une base de données.", [], [], 0, "N/A", "N/A", "N/A"
        except Exception as e:
            return f"Erreur lors du calcul du rapport : {e}", [], [], 0, "N/A", "N/A", "N/A"

    def calculer_total_heures_sortie(self, intervalles_sortie):
        total_heures_sortie = sum(map(lambda x: x[1] - x[0], intervalles_sortie), timedelta())
        return total_heures_sortie


    def calculer_intervalles(self, heures_pointage):
        intervalles_presence = []
        intervalles_sortie = []
        in_time = None
        current_interval = None

        for heure in heures_pointage:
            current_time = datetime.strptime(heure[0], "%Y-%m-%d %H:%M:%S")

            if in_time is None:
                in_time = current_time
            else:
                if current_interval is None:
                    current_interval = (in_time, current_time)
                else:
                    if in_time > current_interval[1]:
                        intervalles_sortie.append(current_interval)
                        current_interval = (in_time, current_time)
                    else:
                        current_interval = (current_interval[0], current_time)

                in_time = None

        if current_interval:
            intervalles_sortie.append(current_interval)

        return intervalles_presence, intervalles_sortie

    def calculer_sum_intervalles_presence(self, intervalles_presence):
        sum_intervalles_presence = sum(map(lambda x: x[1] - x[0], intervalles_presence), timedelta())
        #print(sum_intervalles_presence)
        return str(sum_intervalles_presence)

    def calculer_sum_intervalles_sortie(self, intervalles_sortie):
        sum_intervalles_sortie = sum(map(lambda x: x[1] - x[0], intervalles_sortie), timedelta())
        #print(sum_intervalles_sortie)
        return str(sum_intervalles_sortie)

    def calculer_total_heures_presence(self, intervalles_presence):
        total_heures_presence = sum(map(lambda x: x[1] - x[0], intervalles_presence), timedelta())
        # return str(total_heures_presence)
        return (total_heures_presence)

    def afficher_rapports(self, event=None):
        #personne_saisie = self.personnes.get()  # Personne sélectionnée ou saisie
        personne_saisie = self.combobox_personnes.get()

        try:
            if personne_saisie:
                # Utiliser la personne sélectionnée ou saisie
                personne = self.get_personne_associated_with_input(personne_saisie)
                date = self.date_combobox.get()

                if personne and date:
                    rapport, intervalles_presence, intervalles_sortie, nombre_sorties, sum_intervalles_presence, total_heures_presence, total_heures_sortie = self.calculer_rapport(
                        personne, date)
                    self.result_label.config(text=rapport)
                    self.afficher_treeview_intervalles(self.treeview_intervalles_sortie, intervalles_sortie)

                    if not total_heures_presence:
                        self.label_total_heures_presence.config(text=f"Total des heures de présence : 0")
                    else:
                        heures, reste = divmod(total_heures_presence.seconds, 3600)
                        minutes, secondes = divmod(reste, 60)
                        self.label_total_heures_presence.config(
                            text=f"Total des heures de présence : {heures:02}:{minutes:02}:{secondes:02}")

                        # Afficher total_heures_sortie ici
                    #if not total_heures_sortie:
                    #    self.label_total_heures_sortie.config(text=f"Total des heures de sortie : 0")
                    #else:
                    #    heures, reste = divmod(total_heures_sortie.seconds, 3600)
                    #    minutes, secondes = divmod(reste, 60)
                    #    self.label_total_heures_sortie.config(text=f"Total des heures de sortie : {heures:02}:{minutes:02}:{secondes:02}")

                    self.afficher_treeview_rapports_intermediaires(intervalles_sortie)
                    self.label_nombre_sorties.config(text=f"Nombre de sorties : {nombre_sorties}")
                else:
                    self.result_label.config(text="Veuillez sélectionner ou saisir\n une personne et une date.")

        except Exception as e:
            self.result_label.config(text=f"Erreur : {e}")

    def get_personne_associated_with_input(self, input_str):
        personnes = self.get_personnes()
        suggestions = [p for p in personnes if input_str.lower() in p.lower()]

        if suggestions:
            return suggestions[0]  # Retourner la première suggestion
        else:
            return input_str  # Si aucune suggestion, retourner la saisie

    def calculer_toutes_heures_totales(self):
        try:
            if self.c:
                self.liste_htt = []  # Réinitialiser la liste

                for personne in self.get_personnes():
                    for date in self.get_dates_disponibles():
                        rapport, _, _, _, _, total_heures_presence = self.calculer_rapport(personne, date)
                        # Formatter le total des heures de présence en heures, minutes et secondes
                        if (type(total_heures_presence)==str):
                            self.liste_htt.append({
                                'Noms': personne,
                                'Dates': date,
                                'Total des heures de presence': total_heures_presence
                            })
                        else:
                            heures, reste = divmod(total_heures_presence.seconds, 3600)
                            minutes, secondes = divmod(reste, 60)

                            # Stocker le total des heures de présence formaté pour cette personne et cette date dans la liste
                            self.liste_htt.append({
                                'Noms': personne,
                                'Dates': date,
                                'Total des heures de presencel': f"{heures:02}:{minutes:02}:{secondes:02}"
                            })

                return self.liste_htt

            else:
                return "Aucune base de données importée. Veuillez importer une base de données."

        except Exception as e:
            return f"Erreur lors du calcul des heures totales : {e}"

    def exporter_vers_excel(self):
        try:
            self.liste_htt = self.calculer_toutes_heures_totales()

            if not self.liste_htt:
                messagebox.showinfo("Aucune donnée", "Aucune donnée à exporter.")
                return

            # Créer un DataFrame à partir de la liste
            df = pd.DataFrame(self.liste_htt)

            # Remplacer les valeurs N/A par des durées nulles
            df = df.replace('N/A', pd.Timedelta(seconds=0))

            # Ajouter une colonne avec la somme des valeurs de chaque ligne
            df['Somme'] = df.iloc[:, 2:].apply(lambda row: pd.to_timedelta(row).sum(), axis=1)

            # Créer un tableau croisé dynamique avec pandas
            pivot_table = pd.pivot_table(df, values=['Total des heures de presencel', 'Somme'], index='Noms', columns='Dates',
                                         aggfunc='sum', fill_value=pd.Timedelta(seconds=0))

            # Sélectionner un emplacement de fichier pour l'exportation
            fichier_export = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                          filetypes=[("Fichier Excel", "*.xlsx")])

            if fichier_export:
                # Créer un classeur Excel avec openpyxl
                wb = Workbook()

                # Ajouter une feuille de données au classeur
                ws = wb.active

                # Ajouter le tableau croisé dynamique à la feuille
                for row in dataframe_to_rows(pivot_table, index=True, header=True):
                    ws.append(row)

                # Ajouter un tableau Excel à partir du tableau croisé dynamique
                table = Table(displayName="TableauCroiseDynamique", ref=ws.dimensions)
                style = TableStyleInfo(
                    name="TableStyleMedium9", showFirstColumn=False,
                    showLastColumn=False, showRowStripes=True, showColumnStripes=True)
                table.tableStyleInfo = style
                ws.add_table(table)

                # Enregistrer le classeur Excel
                wb.save(fichier_export)

                messagebox.showinfo("Export réussi",
                                    "Les données ont été exportées avec succès vers un fichier Excel avec un tableau croisé dynamique incluant la somme des valeurs de chaque ligne.")

        except Exception as e:
            messagebox.showerror("Erreur d'exportation", f"Une erreur s'est produite lors de l'exportation : {e}")

    def charger_depuis_excel(self):
        try:
            fichier_excel = filedialog.askopenfilename(title="Sélectionner le fichier Excel",
                                                       filetypes=[("Fichiers Excel", "*.xlsx")])

            if fichier_excel:
                # Lire les données depuis le fichier Excel
                df = pd.read_excel(fichier_excel)

                # Effacer le Treeview existant
                self.treeview_excel_data.delete(*self.treeview_excel_data.get_children())

                # Ajouter les données au Treeview
                for index, row in df.iterrows():
                    self.treeview_excel_data.insert("", index, text=index, values=(
                    row['Noms'], row['Dates'], row['Total des heures de présence']))

        except Exception as e:
            messagebox.showerror("Erreur lors du chargement depuis Excel", f"Une erreur s'est produite : {e}")

    def charger_depuis_excel2(self):
        try:
            fichier_excel = filedialog.askopenfilename(title="Sélectionner le fichier Excel",
                                                       filetypes=[("Fichiers Excel", "*.xlsx")])

            if fichier_excel:
                # Lire les données depuis le fichier Excel
                df = pd.read_excel(fichier_excel)

                # Afficher les données dans une fenêtre séparée
                top = tk.Toplevel(self.root)
                top.title("Données depuis Excel")

                # Créer un Treeview pour afficher les données
                treeview_excel_data = ttk.Treeview(top, columns=df.columns)
                for col in df.columns:
                    treeview_excel_data.heading(col, text=col)
                    treeview_excel_data.column(col, width=150)

                treeview_excel_data.grid(row=0, column=0, padx=10, pady=20, sticky="nw")

                # Ajouter les données au Treeview
                for index, row in df.iterrows():
                    treeview_excel_data.insert("", index, text=index, values=list(row))

        except Exception as e:
            messagebox.showerror("Erreur lors du chargement depuis Excel", f"Une erreur s'est produite : {e}")

    def exporter_intervalles_vers_excel(self):
        try:
            # Obtenir le nom du fichier et le chemin de sortie
            fichier_export = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                          filetypes=[("Fichier Excel", "*.xlsx")])

            if fichier_export:
                # Créer un classeur Excel avec openpyxl
                wb = Workbook()

                # Obtenir toutes les personnes et dates disponibles
                personnes = self.get_personnes()
                dates_disponibles = self.get_dates_disponibles()

                # Ajouter une feuille de données au classeur
                ws = wb.active

                # En-têtes de colonne
                ws.append(["Personne", "Date", "In", "Out"])

                # Ajouter les intervalles des heures de travail au classeur
                for personne in personnes:
                    for date in dates_disponibles:
                        _, _, intervalles_sortie, _, _, _ = self.calculer_rapport(personne, date)
                        for debut, fin in intervalles_sortie:
                            ws.append([personne, date, debut.strftime("%H:%M:%S"), fin.strftime("%H:%M:%S")])

                # Enregistrer le classeur Excel
                wb.save(fichier_export)

                messagebox.showinfo("Export réussi",
                                    "Les intervalles des heures de travail ont été exportés avec succès vers un fichier Excel.")

        except Exception as e:
            messagebox.showerror("Erreur d'exportation", f"Une erreur s'est produite lors de l'exportation : {e}")

class CSVtoDBConverter:
    def __init__(self, master):
        self.master = master
        self.master.title("CSV to DB Converter")

        self.label_source = tk.Label(self.master, text="Chemin du fichier CSV source:")
        self.label_source.grid(row=0, column=0, padx=10, pady=10)
        self.source_var = tk.StringVar()
        self.source_entry = tk.Entry(self.master, textvariable=self.source_var)
        self.source_entry.grid(row=0, column=1, padx=10, pady=10)
        self.button_browse_source = tk.Button(self.master, text="Parcourir", command=self.browse_source)
        self.button_browse_source.grid(row=0, column=2, padx=10, pady=10)

        self.label_destination = tk.Label(self.master, text="Chemin de sortie de la base de données:")
        self.label_destination.grid(row=1, column=0, padx=10, pady=10)
        self.destination_var = tk.StringVar()
        self.destination_entry = tk.Entry(self.master, textvariable=self.destination_var)
        self.destination_entry.grid(row=1, column=1, padx=10, pady=10)
        self.button_browse_destination = tk.Button(self.master, text="Parcourir", command=self.browse_destination)
        self.button_browse_destination.grid(row=1, column=2, padx=10, pady=10)

        self.button_convert = tk.Button(self.master, text="Convertir", command=self.convert_csv_to_db)
        self.button_convert.grid(row=2, column=0, columnspan=3, pady=10)

        self.status_label = tk.Label(self.master, text="")
        self.status_label.grid(row=4, column=0, columnspan=3)

    def browse_source(self):
        file_path = filedialog.askopenfilename(filetypes=[("Fichiers CSV", "*.csv")])
        self.source_var.set(file_path)

    def browse_destination(self):
        folder_path = filedialog.askdirectory()
        self.destination_var.set(folder_path)

    def convert_csv_to_db(self):
        source_path = self.source_var.get()
        destination_path = self.destination_var.get()

        if not source_path or not destination_path:
            self.show_message(False, "Veuillez spécifier le chemin du fichier source et de destination.")
            return

        try:
            conn = sqlite3.connect(destination_path + "/car_pointeuse.db")
            cursor = conn.cursor()

            # Créer la table s'il n'existe pas
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS pointeuse (
                    id_personne INTEGER,
                    nom_personne TEXT,
                    service TEXT,
                    heure TEXT,
                    UNIQUE (id_personne, nom_personne, service, heure)
                )
            ''')

            with open(source_path, 'r', encoding='cp1252') as file:
                csv_reader = csv.reader(file)
                headers = next(csv_reader)

                for row in csv_reader:
                    # Assurez-vous d'ajuster les indices en fonction de votre fichier CSV
                    id_personne = row[0]
                    nom_personne = row[1]
                    service = row[2]
                    heure = row[3]

                    # Insérer les données dans la table (ignorer les doublons)
                    cursor.execute("INSERT OR IGNORE INTO pointeuse VALUES (?, ?, ?, ?)",
                                   (id_personne, nom_personne, service, heure))

            conn.commit()
            conn.close()
            self.show_message(True, "Conversion terminée avec succès.")

        except Exception as e:
            self.show_message(False, f"Une erreur s'est produite : {str(e)}")

    def show_message(self, success, message):
        if success:
            messagebox.showinfo("Succès", message)
        else:
            messagebox.showerror("Erreur", message)

        # Fermer la fenêtre parente
        self.master.destroy()

if __name__ == "__main__":
    root = tk.Tk()
    app = RapportEnregistrementApp(root)
    root.mainloop()
